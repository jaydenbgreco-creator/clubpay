"""
Import members from the uploaded Excel file into MongoDB.
Uses upsert: updates existing members by member_id, inserts new ones.
"""
import requests
import openpyxl
import io
from pymongo import MongoClient
from datetime import datetime, timezone
import uuid

EXCEL_URL = "https://customer-assets.emergentagent.com/job_afterschool-wallet/artifacts/tlkrs4wq_clubbucks_pay_app_final_fixed%20-%20Copy2.xlsx"

client = MongoClient("mongodb://localhost:27017")
db = client["test_database"]

def download_and_parse():
    print("Downloading Excel file...")
    resp = requests.get(EXCEL_URL)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    return wb

def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def import_members(wb):
    ws = wb["Members"]
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h.strip()] = i

    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        member_id = row[col_map.get("Member ID", 0)]
        if not member_id:
            continue

        first_name = row[col_map.get("First Name", 1)] or ""
        last_name = row[col_map.get("Last Name", 2)] or ""
        display = row[col_map.get("Display", 3)] or f"{first_name} {last_name}".strip()
        status = row[col_map.get("Status", 4)] or "Active"
        starting_balance = safe_float(row[col_map.get("Starting Balance", 5)])
        earned = safe_float(row[col_map.get("Earned", 6)])
        bonus = safe_float(row[col_map.get("Bonus", 7)])
        spent = safe_float(row[col_map.get("Spent", 8)])
        adjustments = safe_float(row[col_map.get("Adjustments", 9)])
        current_balance = safe_float(row[col_map.get("Current Balance", 10)])
        qr_payload = row[col_map.get("QR Payload", 11)] or f"CLUBPAY|{member_id}|{display}"
        notes = row[col_map.get("Notes", 12)] or ""

        members.append({
            "member_id": str(member_id).strip(),
            "first_name": str(first_name).strip(),
            "last_name": str(last_name).strip(),
            "display_name": str(display).strip(),
            "status": str(status).strip(),
            "starting_balance": starting_balance,
            "earned": earned,
            "bonus": bonus,
            "spent": spent,
            "adjustments": adjustments,
            "current_balance": current_balance,
            "qr_payload": str(qr_payload).strip(),
            "notes": str(notes).strip(),
        })

    print(f"Parsed {len(members)} members from Excel")
    return members

def upsert_members(members):
    inserted = 0
    updated = 0
    for m in members:
        existing = db.members.find_one({"member_id": m["member_id"]})
        if existing:
            # Update fields but preserve existing balances if they have activity
            update_fields = {
                "first_name": m["first_name"],
                "last_name": m["last_name"],
                "display_name": m["display_name"],
                "status": m["status"],
                "qr_payload": m["qr_payload"],
                "notes": m["notes"],
            }
            # Only update balance fields if the existing record has zero activity
            if existing.get("earned", 0) == 0 and existing.get("spent", 0) == 0 and existing.get("bonus", 0) == 0 and existing.get("adjustments", 0) == 0:
                update_fields["starting_balance"] = m["starting_balance"]
                update_fields["earned"] = m["earned"]
                update_fields["bonus"] = m["bonus"]
                update_fields["spent"] = m["spent"]
                update_fields["adjustments"] = m["adjustments"]
                update_fields["current_balance"] = m["current_balance"]

            db.members.update_one({"member_id": m["member_id"]}, {"$set": update_fields})
            updated += 1
        else:
            m["id"] = str(uuid.uuid4())
            m["created_at"] = datetime.now(timezone.utc).isoformat()
            db.members.insert_one(m)
            inserted += 1

    print(f"Import complete: {inserted} new members inserted, {updated} existing members updated")
    return inserted, updated

if __name__ == "__main__":
    wb = download_and_parse()
    members = import_members(wb)
    inserted, updated = upsert_members(members)
    
    # Verify
    total = db.members.count_documents({})
    print(f"Total members in database: {total}")
